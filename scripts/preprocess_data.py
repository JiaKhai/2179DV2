from __future__ import annotations

import csv
import datetime as dt
import json
import math
import re
import shutil
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "data" / "processed"
PUBLIC_OUT_DIR = ROOT / "public" / "data" / "processed"


RAW_FLOOD = RAW_DIR / "SPECIAL REPORT ON IMPACT OF FLOODS IN MALAYSIA, 2025.xlsx"
RAW_POPULATION = RAW_DIR / "population_state.csv"
RAW_RAINFALL = RAW_DIR / "purata-hujan-tahunan-2013-2017.xlsx"
RAW_STATIONS = RAW_DIR / "stesen-hujan-di-malaysia.xlsx"
RAW_GEOJSON = RAW_DIR / "MYS_ADM1.json"


STATE_ALIASES = {
    "JOHOR": "Johor",
    "KEDAH": "Kedah",
    "KELANTAN": "Kelantan",
    "MELAKA": "Melaka",
    "MALACCA": "Melaka",
    "NEGERI SEMBILAN": "Negeri Sembilan",
    "PAHANG": "Pahang",
    "PERAK": "Perak",
    "PERLIS": "Perlis",
    "PULAU PINANG": "Pulau Pinang",
    "PENANG": "Pulau Pinang",
    "SABAH": "Sabah",
    "SARAWAK": "Sarawak",
    "SELANGOR": "Selangor",
    "TERENGGANU": "Terengganu",
    "KUALA LUMPUR": "W.P. Kuala Lumpur",
    "WP KUALA LUMPUR": "W.P. Kuala Lumpur",
    "W.P. KUALA LUMPUR": "W.P. Kuala Lumpur",
    "WILAYAH PERSEKUTUAN KUALA LUMPUR": "W.P. Kuala Lumpur",
    "LABUAN": "W.P. Labuan",
    "W.P. LABUAN": "W.P. Labuan",
    "PUTRAJAYA": "W.P. Putrajaya",
    "W.P. PUTRAJAYA": "W.P. Putrajaya",
    "MALAYSIA": "Malaysia",
}


GEO_ISO = {
    "Johor": "MY-01",
    "Kedah": "MY-02",
    "Kelantan": "MY-03",
    "Melaka": "MY-04",
    "Negeri Sembilan": "MY-05",
    "Pahang": "MY-06",
    "Pulau Pinang": "MY-07",
    "Perak": "MY-08",
    "Perlis": "MY-09",
    "Selangor": "MY-10",
    "Terengganu": "MY-11",
    "Sabah": "MY-12",
    "Sarawak": "MY-13",
    "W.P. Kuala Lumpur": "MY-14",
    "W.P. Labuan": "MY-15",
    "W.P. Putrajaya": "MY-16",
}


STATE_REGION = {
    "Johor": "Southern Peninsula",
    "Kedah": "Northern Peninsula",
    "Kelantan": "East Coast",
    "Melaka": "Southern Peninsula",
    "Negeri Sembilan": "Central/Southern Peninsula",
    "Pahang": "East Coast",
    "Perak": "Northern Peninsula",
    "Perlis": "Northern Peninsula",
    "Pulau Pinang": "Northern Peninsula",
    "Sabah": "Borneo",
    "Sarawak": "Borneo",
    "Selangor": "Central Peninsula",
    "Terengganu": "East Coast",
    "W.P. Kuala Lumpur": "Central Peninsula",
    "W.P. Labuan": "Borneo",
    "W.P. Putrajaya": "Central Peninsula",
}


def canonical_state(value: Any) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value).strip())
    key = text.upper().replace(".", "")
    key = key.replace("W P ", "WP ")
    return STATE_ALIASES.get(key, text.title())


def numeric(value: Any) -> float | None:
    if value in (None, "", "#REF!"):
        return None
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        num = float(value)
        if math.isnan(num):
            return None
        return num
    except (TypeError, ValueError):
        return None


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def fmt(value: Any, digits: int = 3) -> Any:
    if isinstance(value, float):
        return round(value, digits)
    return value


def clean_record(record: dict[str, Any]) -> dict[str, Any]:
    return {key: fmt(value) for key, value in record.items()}


def preprocess_state_lookup() -> list[dict[str, str]]:
    rows = []
    for state, iso in GEO_ISO.items():
        rows.append(
            {
                "state": state,
                "iso_code": iso,
                "region": STATE_REGION[state],
            }
        )
    write_csv(OUT_DIR / "state_lookup.csv", rows, ["state", "iso_code", "region"])
    return rows


def preprocess_flood_state() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(RAW_FLOOD, read_only=True, data_only=True)
    ws = wb["Jadual 1"]
    allowed_states = set(GEO_ISO) | {"Malaysia"}

    cols = {
        "living_quarters_loss_rm000": {2024: 6, 2025: 7},
        "vehicles_loss_rm000": {2024: 9, 2025: 10},
        "business_premises_loss_rm000": {2024: 12, 2025: 13},
        "manufacturing_loss_rm000": {2024: 15, 2025: 16},
        "agriculture_loss_rm000": {2024: 18, 2025: 19},
        "infrastructure_loss_rm000": {2024: 21, 2025: 22},
        "total_loss_rm000": {2024: 24, 2025: 25},
    }

    rows: list[dict[str, Any]] = []
    for row_idx in range(9, ws.max_row + 1):
        state = canonical_state(ws.cell(row_idx, 2).value)
        if state not in allowed_states:
            continue
        living_quarters_count = numeric(ws.cell(row_idx, 4).value)
        for year in [2024, 2025]:
            rec: dict[str, Any] = {
                "state": state,
                "year": year,
                "living_quarters_count": int(living_quarters_count) if living_quarters_count is not None else None,
            }
            for field, year_cols in cols.items():
                rec[field] = numeric(ws.cell(row_idx, year_cols[year]).value) or 0
            rec["total_loss_rm_million"] = rec["total_loss_rm000"] / 1000
            rows.append(clean_record(rec))

    fieldnames = [
        "state",
        "year",
        "living_quarters_count",
        "living_quarters_loss_rm000",
        "vehicles_loss_rm000",
        "business_premises_loss_rm000",
        "manufacturing_loss_rm000",
        "agriculture_loss_rm000",
        "infrastructure_loss_rm000",
        "total_loss_rm000",
        "total_loss_rm_million",
    ]
    write_csv(OUT_DIR / "flood_losses_state.csv", rows, fieldnames)
    return rows


def preprocess_flood_district() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(RAW_FLOOD, read_only=True, data_only=True)
    ws = wb["Jadual 2"]

    rows: list[dict[str, Any]] = []
    current_state = ""
    state_names = set(GEO_ISO) | {"Malaysia"}
    for row_idx in range(8, ws.max_row + 1):
        label_raw = ws.cell(row_idx, 3).value
        label = canonical_state(label_raw)
        if not label:
            continue

        if label in state_names:
            current_state = label
            place_type = "state"
            district = ""
        else:
            place_type = "district"
            district = str(label_raw).strip()

        living_quarters_count = numeric(ws.cell(row_idx, 5).value)
        for year, total_col in [(2024, 19), (2025, 20)]:
            total_loss = numeric(ws.cell(row_idx, total_col).value) or 0
            rows.append(
                clean_record(
                    {
                        "state": current_state,
                        "district": district,
                        "place_name": current_state if place_type == "state" else district,
                        "place_type": place_type,
                        "year": year,
                        "living_quarters_count": int(living_quarters_count) if living_quarters_count is not None else None,
                        "total_loss_rm000": total_loss,
                        "total_loss_rm_million": total_loss / 1000,
                    }
                )
            )

    fieldnames = [
        "state",
        "district",
        "place_name",
        "place_type",
        "year",
        "living_quarters_count",
        "total_loss_rm000",
        "total_loss_rm_million",
    ]
    write_csv(OUT_DIR / "flood_losses_district.csv", rows, fieldnames)
    return rows


def preprocess_population() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with RAW_POPULATION.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["sex"] == "both" and row["age"] == "overall" and row["ethnicity"] == "overall":
                year = int(row["date"][:4])
                state = canonical_state(row["state"])
                rows.append(
                    {
                        "state": state,
                        "year": year,
                        "population_thousand": numeric(row["population"]),
                        "population": int(round((numeric(row["population"]) or 0) * 1000)),
                    }
                )

    write_csv(OUT_DIR / "population_state_overall.csv", rows, ["state", "year", "population_thousand", "population"])
    latest_year = max(int(row["year"]) for row in rows)
    latest = [row for row in rows if int(row["year"]) == latest_year]
    write_csv(OUT_DIR / "population_state_latest.csv", latest, ["state", "year", "population_thousand", "population"])
    return rows


def preprocess_rainfall() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(RAW_RAINFALL, read_only=True, data_only=True)
    ws = wb.active
    years = [int(ws.cell(5, col).value) for col in range(3, 8)]
    rows: list[dict[str, Any]] = []

    for row_idx in range(6, ws.max_row + 1):
        state = canonical_state(ws.cell(row_idx, 2).value)
        if not state:
            continue
        values: list[float] = []
        for col, year in zip(range(3, 8), years):
            rainfall = numeric(ws.cell(row_idx, col).value)
            if rainfall is None:
                continue
            values.append(rainfall)
            rows.append(
                {
                    "state": state,
                    "year": year,
                    "annual_rainfall_mm": round(rainfall, 1),
                }
            )
        if values:
            avg = sum(values) / len(values)
            rows.append(
                {
                    "state": state,
                    "year": "2013-2017 average",
                    "annual_rainfall_mm": round(avg, 1),
                }
            )

    write_csv(OUT_DIR / "rainfall_state_annual.csv", rows, ["state", "year", "annual_rainfall_mm"])
    return rows


def parse_dms_text(value: str, axis: str) -> tuple[float | None, str]:
    text = value.strip().strip("'").strip('"')
    parts = [p.strip(".") for p in re.split(r"[^0-9.]+", text) if p.strip(".") != ""]
    if not parts:
        return None, "missing"
    if len(parts) >= 3:
        deg, minutes, seconds = map(float, parts[:3])
        return deg + minutes / 60 + seconds / 3600, "parsed_dms"
    if len(parts) == 2:
        deg, minutes = map(float, parts)
        return deg + minutes / 60, "parsed_dm"
    return parse_numeric_coord(float(parts[0]), axis)


def parse_numeric_coord(value: float, axis: str) -> tuple[float | None, str]:
    abs_value = abs(value)
    max_decimal = 8 if axis == "lat" else 120
    if abs_value <= max_decimal:
        return value, "already_decimal"

    whole = int(abs_value)
    frac = abs_value - whole
    if axis == "lat":
        deg = whole // 100
        minutes = (whole % 100) + frac
    else:
        deg = whole // 100
        minutes = (whole % 100) + frac
    if minutes >= 60:
        return None, "invalid_decimal_minutes"
    decimal = deg + minutes / 60
    return (-decimal if value < 0 else decimal), "parsed_decimal_minutes"


def parse_coord(value: Any, axis: str) -> tuple[float | None, str]:
    if value in (None, ""):
        return None, "missing"
    if isinstance(value, dt.datetime):
        deg = value.month
        minutes = value.day
        seconds = value.year % 100
        return deg + minutes / 60 + seconds / 3600, "parsed_excel_date"
    if isinstance(value, dt.date):
        deg = value.month
        minutes = value.day
        seconds = value.year % 100
        return deg + minutes / 60 + seconds / 3600, "parsed_excel_date"
    if isinstance(value, (int, float)):
        return parse_numeric_coord(float(value), axis)
    return parse_dms_text(str(value), axis)


def preprocess_stations() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(RAW_STATIONS, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    seen = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        station_no, station_name, function, state, raw_lat, raw_lon = row[:6]
        if not station_no or station_no == "Note :":
            continue
        lat, lat_status = parse_coord(raw_lat, "lat")
        lon, lon_status = parse_coord(raw_lon, "lon")
        coord_status = "ok" if lat is not None and lon is not None else "missing"
        if coord_status == "ok" and not (0 <= lat <= 8 and 99 <= lon <= 120):
            coord_status = "outside_malaysia_bounds"
        key = (station_no, station_name, state, lat, lon)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "station_no": str(station_no),
                "station_name": str(station_name).strip(),
                "function": str(function).strip() if function else "",
                "state_raw": str(state).strip() if state else "",
                "state": canonical_state(state) if state else "",
                "latitude": round(lat, 6) if lat is not None else "",
                "longitude": round(lon, 6) if lon is not None else "",
                "coord_status": coord_status,
                "lat_parse_method": lat_status,
                "lon_parse_method": lon_status,
            }
        )

    fieldnames = [
        "station_no",
        "station_name",
        "function",
        "state_raw",
        "state",
        "latitude",
        "longitude",
        "coord_status",
        "lat_parse_method",
        "lon_parse_method",
    ]
    write_csv(OUT_DIR / "rainfall_stations_clean.csv", rows, fieldnames)
    usable = [row for row in rows if row["coord_status"] == "ok"]
    write_csv(OUT_DIR / "rainfall_stations_mappable.csv", usable, fieldnames)
    return rows


def round_geometry(value: Any, digits: int = 5) -> Any:
    if isinstance(value, list):
        return [round_geometry(item, digits) for item in value]
    if isinstance(value, float):
        return round(value, digits)
    return value


def rewind_geometry(geometry: dict[str, Any]) -> None:
    # Vega/d3 can render some boundary sources as polygon complements when ring
    # winding is opposite of the expected spherical orientation. Reversing rings
    # keeps the coordinates intact but makes Malaysia render as land, not world.
    if geometry.get("type") == "Polygon":
        geometry["coordinates"] = [list(reversed(ring)) for ring in geometry["coordinates"]]
    elif geometry.get("type") == "MultiPolygon":
        geometry["coordinates"] = [
            [list(reversed(ring)) for ring in polygon]
            for polygon in geometry["coordinates"]
        ]


def preprocess_geojson() -> dict[str, Any]:
    geo = json.loads(RAW_GEOJSON.read_text(encoding="utf-8"))
    for feature in geo["features"]:
        raw_name = feature["properties"].get("NAME", "")
        state = canonical_state(raw_name)
        feature["properties"] = {
            "state": state,
            "iso_code": GEO_ISO.get(state, feature["properties"].get("ISO_Code", "")),
            "region": STATE_REGION.get(state, ""),
        }
        rewind_geometry(feature["geometry"])
        feature["geometry"]["coordinates"] = round_geometry(feature["geometry"]["coordinates"], 5)
    output = OUT_DIR / "malaysia_adm1.geojson"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(geo, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return geo


def preprocess_story_state_summary(
    flood_rows: list[dict[str, Any]], population_rows: list[dict[str, Any]], rainfall_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    population_2025 = {row["state"]: row for row in population_rows if row["year"] == 2025}
    rainfall_avg = {
        row["state"]: row["annual_rainfall_mm"] for row in rainfall_rows if row["year"] == "2013-2017 average"
    }
    flood_2025 = {row["state"]: row for row in flood_rows if row["year"] == 2025}
    rows: list[dict[str, Any]] = []

    for state in GEO_ISO:
        flood = flood_2025.get(state, {})
        population = population_2025.get(state, {})
        total_loss = float(flood.get("total_loss_rm000") or 0)
        pop = int(population.get("population") or 0)
        rows.append(
            clean_record(
                {
                    "state": state,
                    "iso_code": GEO_ISO[state],
                    "region": STATE_REGION[state],
                    "population_2025": pop,
                    "annual_rainfall_avg_2013_2017_mm": rainfall_avg.get(state, ""),
                    "flood_total_loss_2025_rm000": total_loss,
                    "flood_total_loss_2025_rm_million": total_loss / 1000,
                    "flood_loss_2025_rm_per_person": (total_loss * 1000 / pop) if pop else "",
                    "living_quarters_count_2025": flood.get("living_quarters_count", ""),
                }
            )
        )

    fieldnames = [
        "state",
        "iso_code",
        "region",
        "population_2025",
        "annual_rainfall_avg_2013_2017_mm",
        "flood_total_loss_2025_rm000",
        "flood_total_loss_2025_rm_million",
        "flood_loss_2025_rm_per_person",
        "living_quarters_count_2025",
    ]
    write_csv(OUT_DIR / "story_state_summary.csv", rows, fieldnames)
    return rows


def mirror_processed_for_vite() -> None:
    PUBLIC_OUT_DIR.mkdir(parents=True, exist_ok=True)
    for source in OUT_DIR.iterdir():
        if source.is_file() and source.suffix.lower() in {".csv", ".json", ".geojson"}:
            shutil.copy2(source, PUBLIC_OUT_DIR / source.name)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    preprocess_state_lookup()
    flood_state = preprocess_flood_state()
    preprocess_flood_district()
    population = preprocess_population()
    rainfall = preprocess_rainfall()
    preprocess_stations()
    preprocess_geojson()
    preprocess_story_state_summary(flood_state, population, rainfall)
    mirror_processed_for_vite()
    print(f"Processed data written to {OUT_DIR}")
    print(f"Public data mirrored to {PUBLIC_OUT_DIR}")


if __name__ == "__main__":
    main()
